"""OMNIA — AI Smart Import Clienti (D-FUTURE-07, v1).

Pattern: file disordinato (CSV/Excel/vCard/TXT) → Gemini-3-flash structured output
→ draft con preview + confidence per riga → conferma + commit batch.

Endpoints (mounted under /app/clients/import/ai):
  POST /                      → upload file + parse + create draft (TTL 1h)
  GET  /draft/{id}            → load draft for preview
  PATCH /draft/{id}/row/{idx} → edit a single row
  POST /draft/{id}/commit     → finalize import (insert clients)

Future v2 (D-FUTURE-09): PDF + Gemini Vision on screenshots.
"""
import asyncio
import csv
import io
import json
import logging
import os
import re
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel, Field

from shared.auth.dependencies import get_current_user, require_roles
from shared.db.connection import Database

logger = logging.getLogger("omnia.clients_ai_import")
router = APIRouter(prefix="/clients/import/ai", tags=["clients-ai-import"])


# ============================================================
# CONFIG
# ============================================================

MAX_FILE_BYTES = 5 * 1024 * 1024          # 5 MB
MAX_ROWS_PER_IMPORT = 500
DRAFT_TTL_MINUTES = 60
GEMINI_MODEL = "gemini-3-flash-preview"
GEMINI_BATCH_SIZE = 25                     # rows per LLM call (cap context size)

CLIENT_TYPES = ("buyer", "seller", "tenant", "landlord", "investor")
OPERATIONS = ("sale", "rent")
PROPERTY_TYPES = (
    "apartment", "house", "villa", "townhouse", "loft", "penthouse",
    "studio", "duplex", "room", "garage", "land", "office",
    "shop", "warehouse", "building", "other",
)


# ============================================================
# Gemini prompt
# ============================================================

SYSTEM_PROMPT = """Sei un data analyst esperto in immobiliare italiano.
Ricevi una porzione disordinata di dati clienti (può essere righe di CSV/Excel con colonne arbitrarie, contatti vCard, o testo libero con note d'agenzia).
Devi mappare i clienti allo schema OMNIA, una riga per cliente reale.

REGOLE FERREE:
- Rispondi SOLO con un array JSON valido. Niente testo prima/dopo. Niente backtick markdown.
- Lingua: italiano per i campi `notes`.
- Una riga di output = un cliente reale. Salta righe che sono header, separatori, note generiche.
- Se un campo non è deducibile metti null (NON inventare).
- `confidence` (0-100): quanto sei sicuro che la riga sia un cliente valido (90+ se ci sono name + (email o phone), 50-89 se manca qualcosa, <50 se molto incerto).
- `warnings`: array di stringhe brevi in italiano sui campi dubbi/mancanti.

SCHEMA per ogni elemento dell'array:
{
  "name": "string (REQUIRED, mai vuoto)",
  "surname": "string o null",
  "email": "email valida o null",
  "phone": "stringa numerica (può includere + e spazi) o null",
  "whatsapp": "stringa numerica o null (se diverso da phone)",
  "client_type": "buyer | seller | tenant | landlord | investor (deduci dal contesto: 'cerca' → buyer/tenant, 'vende/affitta' → seller/landlord, 'investe' → investor; default 'buyer' se ambiguo)",
  "status": "new | qualified | active | archived (default 'new' se non specificato)",
  "preferences": {
    "operation": "sale | rent o null",
    "property_types": ["apartment" | "house" | ...] o [],
    "cities": ["Roma", "Milano", ...] o [],
    "zones": [...] o [],
    "price_max": numero o null,
    "price_min": numero o null,
    "surface_min": numero o null,
    "rooms_min": numero o null,
    "bedrooms_min": numero o null
  },
  "notes": "stringa concisa con info residue (max 300 caratteri) o null",
  "confidence": 0-100,
  "warnings": ["stringa", ...]
}

ESEMPI di interpretazione intelligente:
- "tel 333..." o "cell" o "mobile" → phone
- "città" / "zona richiesta" / "cerca a" / "interessato a" → preferences.cities
- "budget", "max", "fino a", "<300k" → preferences.price_max
- "trilocale", "2 stanze", "T3" → preferences.rooms_min
- "venditore", "ha incarico", "in vendita" → client_type = seller
- "cerca", "acquirente", "compratore" → client_type = buyer
- "appartamento", "casa", "villa" → preferences.property_types

REGOLA FINALE: se ricevi righe che chiaramente NON sono clienti (header, separatori, footer, righe vuote), NON includerle nell'array. Meglio meno righe ma giuste."""


# ============================================================
# Pre-parsers (format-specific extraction → normalized text)
# ============================================================

def _parse_csv(content: bytes) -> List[Dict[str, str]]:
    text = content.decode("utf-8-sig", errors="replace")
    # Detect separator
    sample = text[:2000]
    sep = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=sep)
    rows = []
    for r in reader:
        # strip + dedupe whitespace; convert None keys (extra columns) to "_extra"
        cleaned = {(k or "_extra").strip(): (v or "").strip() for k, v in r.items()}
        if any(cleaned.values()):
            rows.append(cleaned)
    return rows[:MAX_ROWS_PER_IMPORT]


def _parse_xlsx(content: bytes) -> List[Dict[str, str]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h or f"col_{i}").strip() for i, h in enumerate(next(rows_iter))]
    except StopIteration:
        return []
    out = []
    for row in rows_iter:
        if all(v in (None, "") for v in row):
            continue
        rec = {}
        for i, val in enumerate(row):
            if i >= len(headers):
                break
            rec[headers[i]] = "" if val is None else str(val).strip()
        out.append(rec)
        if len(out) >= MAX_ROWS_PER_IMPORT:
            break
    return out


def _parse_vcard(content: bytes) -> List[Dict[str, str]]:
    import vobject
    text = content.decode("utf-8", errors="replace")
    out = []
    for v in vobject.readComponents(text):
        try:
            fn = (v.fn.value if hasattr(v, "fn") else "") or ""
            tel = []
            for t in v.contents.get("tel", []) or []:
                tel.append(getattr(t, "value", ""))
            email = ""
            if "email" in v.contents:
                email = v.contents["email"][0].value
            note = ""
            if "note" in v.contents:
                note = v.contents["note"][0].value
            org = ""
            if "org" in v.contents:
                org_val = v.contents["org"][0].value
                org = " ".join(org_val) if isinstance(org_val, list) else str(org_val)
            out.append({
                "fn": fn,
                "tel": "; ".join(tel),
                "email": email,
                "note": note,
                "org": org,
            })
        except Exception:
            continue
        if len(out) >= MAX_ROWS_PER_IMPORT:
            break
    return out


def _parse_text(content: bytes) -> List[Dict[str, str]]:
    """For .txt files we don't pre-structure. Send line groups to Gemini as-is."""
    text = content.decode("utf-8", errors="replace")
    # split by blank lines; each chunk is one "row" for Gemini
    chunks = [c.strip() for c in re.split(r"\n\s*\n", text) if c.strip()]
    rows = [{"_raw": c} for c in chunks[:MAX_ROWS_PER_IMPORT]]
    if not rows:
        # fallback: each non-empty line is a row
        rows = [{"_raw": ln.strip()} for ln in text.splitlines() if ln.strip()][:MAX_ROWS_PER_IMPORT]
    return rows


_FORMAT_PARSERS = {
    "csv": _parse_csv,
    "xlsx": _parse_xlsx,
    "vcf": _parse_vcard,
    "txt": _parse_text,
}


def _detect_format(filename: str, content_type: Optional[str], head_bytes: bytes) -> str:
    name = (filename or "").lower().strip()
    ext = name.rsplit(".", 1)[-1] if "." in name else ""
    if ext in _FORMAT_PARSERS:
        return ext
    head = head_bytes[:200].decode("utf-8", errors="ignore").lower()
    if "begin:vcard" in head:
        return "vcf"
    if head_bytes[:4] == b"PK\x03\x04":   # ZIP magic → xlsx
        return "xlsx"
    if "," in head or ";" in head:
        return "csv"
    return "txt"


# ============================================================
# Gemini batch caller
# ============================================================

async def _gemini_extract_batch(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Send a chunk of pre-parsed rows to Gemini → return mapped client array."""
    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(status_code=503, detail="emergent_llm_key_missing")
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage  # type: ignore
        chat = LlmChat(
            api_key=api_key,
            session_id=f"client-ai-import-{uuid4()}",
            system_message=SYSTEM_PROMPT,
        ).with_model("gemini", GEMINI_MODEL)
        # Truncate each row representation to keep prompt bounded
        payload = json.dumps(rows, ensure_ascii=False)[:18000]
        user_text = "Mappa i seguenti record clienti allo schema OMNIA. Rispondi con SOLO l'array JSON:\n\n" + payload
        raw = await chat.send_message(UserMessage(text=user_text))
        text = (raw or "").strip()
        # strip markdown code fences if present
        if text.startswith("```"):
            text = text.split("```", 2)[1] if text.count("```") >= 2 else text
            if text.lower().startswith("json"):
                text = text[4:]
        text = text.strip()
        if not text.startswith("["):
            # Try to locate the array
            m = re.search(r"\[\s*\{.*\}\s*\]", text, re.DOTALL)
            if m:
                text = m.group(0)
        return json.loads(text)
    except json.JSONDecodeError as e:
        logger.warning(f"Gemini AI import JSON decode failed: {e}")
        return []
    except Exception as e:
        logger.warning(f"AI import extraction failed: {type(e).__name__}: {e}")
        return []


# ============================================================
# Row validation / sanitization (defensive layer over Gemini output)
# ============================================================

def _sanitize_email(e: Any) -> Optional[str]:
    if not e:
        return None
    s = str(e).strip()
    if "@" not in s or " " in s:
        return None
    return s.lower()


def _sanitize_phone(p: Any) -> Optional[str]:
    if not p:
        return None
    s = re.sub(r"[^\d+]", "", str(p))
    if len(re.sub(r"\D", "", s)) < 6:
        return None
    return s


def _coerce_enum(v: Any, allowed: tuple, default: Optional[str]) -> Optional[str]:
    if not v:
        return default
    s = str(v).strip().lower()
    return s if s in allowed else default


def _coerce_int(v: Any) -> Optional[int]:
    if v in (None, "", "null"):
        return None
    try:
        return int(float(str(v).replace(".", "").replace(",", ".")))
    except (ValueError, TypeError):
        return None


def _coerce_list(v: Any, allowed: Optional[tuple] = None) -> List[str]:
    if not v:
        return []
    if isinstance(v, list):
        items = [str(x).strip() for x in v if x]
    elif isinstance(v, str):
        items = [x.strip() for x in re.split(r"[;,]", v) if x.strip()]
    else:
        items = [str(v).strip()]
    if allowed:
        items = [i for i in items if i.lower() in allowed]
    return items


def _normalize_row(r: Dict[str, Any]) -> Dict[str, Any]:
    """Apply schema validation + sanitization. Returns a clean row dict
    or {} if the row is unsalvageable (no name)."""
    name = (r.get("name") or "").strip()
    if not name:
        return {}
    prefs_raw = r.get("preferences") or {}
    prefs = {
        "operation": _coerce_enum(prefs_raw.get("operation"), OPERATIONS, None),
        "property_types": _coerce_list(prefs_raw.get("property_types"), PROPERTY_TYPES),
        "cities": _coerce_list(prefs_raw.get("cities")),
        "zones": _coerce_list(prefs_raw.get("zones")),
        "price_max": _coerce_int(prefs_raw.get("price_max")),
        "price_min": _coerce_int(prefs_raw.get("price_min")),
        "surface_min": _coerce_int(prefs_raw.get("surface_min")),
        "rooms_min": _coerce_int(prefs_raw.get("rooms_min")),
        "bedrooms_min": _coerce_int(prefs_raw.get("bedrooms_min")),
    }
    confidence_raw = r.get("confidence")
    try:
        confidence = max(0, min(100, int(confidence_raw)))
    except (TypeError, ValueError):
        confidence = 50
    return {
        "name": name[:120],
        "surname": ((r.get("surname") or "").strip() or None),
        "email": _sanitize_email(r.get("email")),
        "phone": _sanitize_phone(r.get("phone")),
        "whatsapp": _sanitize_phone(r.get("whatsapp")),
        "client_type": _coerce_enum(r.get("client_type"), CLIENT_TYPES, "buyer"),
        "status": _coerce_enum(r.get("status"), ("new", "qualified", "active", "archived"), "new"),
        "preferences": prefs,
        "notes": ((r.get("notes") or "")[:500] or None),
        "confidence": confidence,
        "warnings": list(r.get("warnings") or [])[:10],
    }


# ============================================================
# Helpers
# ============================================================

from shared.auth.tenant import arequire_agency as _agency_id


def _draft_doc(agency_id: str, user_id: str, filename: str, fmt: str,
               size: int, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    now = datetime.now(timezone.utc)
    return {
        "id": str(uuid4()),
        "agency_id": agency_id,
        "created_by": user_id,
        "created_at": now,
        "expires_at": now + timedelta(minutes=DRAFT_TTL_MINUTES),
        "source_filename": filename,
        "source_format": fmt,
        "source_size_bytes": size,
        "engine": GEMINI_MODEL,
        "rows": rows,
        "status": "draft",
    }


async def _ensure_ttl_index():
    db = Database.get()
    try:
        await db.ai_client_import_drafts.create_index(
            "expires_at", expireAfterSeconds=0, background=True,
        )
    except Exception:
        pass


# ============================================================
# 1) Upload + parse + draft
# ============================================================

@router.post("")
async def upload_and_parse(
    file: UploadFile = File(...),
    user: dict = Depends(require_roles("agency_admin", "agent", "super_admin")),
):
    """Upload a messy clients file (CSV/Excel/vCard/TXT) → Gemini parse → draft preview."""
    await _ensure_ttl_index()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="empty_file")
    if len(content) > MAX_FILE_BYTES:
        raise HTTPException(status_code=413, detail=f"file_too_large_max_{MAX_FILE_BYTES}")

    fmt = _detect_format(file.filename or "", file.content_type, content)
    parser = _FORMAT_PARSERS.get(fmt)
    if not parser:
        raise HTTPException(status_code=415, detail=f"format_not_supported:{fmt}")

    try:
        pre_rows = parser(content)
    except Exception as e:
        logger.warning("pre-parser failed: %s", e)
        raise HTTPException(status_code=400, detail=f"parse_failed:{type(e).__name__}")

    if not pre_rows:
        raise HTTPException(status_code=400, detail="no_rows_detected")

    # Batch via Gemini
    batches = [pre_rows[i:i + GEMINI_BATCH_SIZE]
               for i in range(0, len(pre_rows), GEMINI_BATCH_SIZE)]
    batch_results = await asyncio.gather(*(_gemini_extract_batch(b) for b in batches))
    extracted: List[Dict[str, Any]] = []
    for b in batch_results:
        if isinstance(b, list):
            extracted.extend(b)

    # Normalize + drop empty rows + attach idx + a raw excerpt for the UI
    rows_out: List[Dict[str, Any]] = []
    for idx, raw in enumerate(extracted):
        clean = _normalize_row(raw if isinstance(raw, dict) else {})
        if not clean:
            continue
        # Try to attach a small excerpt of the original line(s) for the preview UI
        clean["idx"] = len(rows_out)
        clean["source_excerpt"] = _make_excerpt(pre_rows, idx)
        rows_out.append(clean)

    if not rows_out:
        raise HTTPException(status_code=422, detail="ai_extracted_no_valid_rows")

    # Persist draft
    agency_id = await _agency_id(user)
    db = Database.get()
    doc = _draft_doc(
        agency_id=agency_id, user_id=user.get("id") or "unknown",
        filename=file.filename or "upload", fmt=fmt,
        size=len(content), rows=rows_out,
    )
    await db.ai_client_import_drafts.insert_one(doc)

    return _draft_response(doc)


def _make_excerpt(pre_rows: List[Dict[str, Any]], idx: int) -> str:
    if 0 <= idx < len(pre_rows):
        r = pre_rows[idx]
        if "_raw" in r:
            return str(r["_raw"])[:240]
        return " · ".join(f"{k}={v}" for k, v in r.items() if v)[:240]
    return ""


def _draft_response(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Build the JSON response sent to the client (omit internal Mongo `_id`)."""
    rows = doc.get("rows") or []
    high = sum(1 for r in rows if r.get("confidence", 0) >= 80)
    medium = sum(1 for r in rows if 50 <= r.get("confidence", 0) < 80)
    low = sum(1 for r in rows if r.get("confidence", 0) < 50)
    return {
        "draft_id": doc["id"],
        "source_filename": doc.get("source_filename"),
        "source_format": doc.get("source_format"),
        "engine": doc.get("engine"),
        "total_rows": len(rows),
        "confidence_buckets": {"high": high, "medium": medium, "low": low},
        "expires_at": doc["expires_at"].isoformat() if isinstance(doc.get("expires_at"), datetime) else doc.get("expires_at"),
        "rows": rows,
    }


# ============================================================
# 2) Get draft for preview (re-load on refresh)
# ============================================================

@router.get("/draft/{draft_id}")
async def get_draft(
    draft_id: str,
    user: dict = Depends(get_current_user),
):
    db = Database.get()
    doc = await db.ai_client_import_drafts.find_one({"id": draft_id})
    if not doc:
        raise HTTPException(status_code=404, detail="draft_not_found")
    agency_id = await _agency_id(user)
    if doc.get("agency_id") != agency_id:
        raise HTTPException(status_code=403, detail="not_your_draft")
    return _draft_response(doc)


# ============================================================
# 3) Edit a single row before commit
# ============================================================

class RowPatch(BaseModel):
    name: Optional[str] = None
    surname: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    whatsapp: Optional[str] = None
    client_type: Optional[str] = None
    status: Optional[str] = None
    preferences: Optional[Dict[str, Any]] = None
    notes: Optional[str] = None
    drop: bool = Field(default=False, description="If true, mark the row to be excluded from commit.")


@router.patch("/draft/{draft_id}/row/{idx}")
async def patch_row(
    draft_id: str,
    idx: int,
    patch: RowPatch,
    user: dict = Depends(require_roles("agency_admin", "agent", "super_admin")),
):
    db = Database.get()
    doc = await db.ai_client_import_drafts.find_one({"id": draft_id})
    if not doc:
        raise HTTPException(status_code=404, detail="draft_not_found")
    agency_id = await _agency_id(user)
    if doc.get("agency_id") != agency_id:
        raise HTTPException(status_code=403, detail="not_your_draft")
    rows = doc.get("rows") or []
    if idx < 0 or idx >= len(rows):
        raise HTTPException(status_code=404, detail="row_not_found")

    row = rows[idx]
    if patch.drop:
        row["_drop"] = True
    else:
        row.pop("_drop", None)
        upd = patch.model_dump(exclude_none=True, exclude={"drop"})
        if upd:
            row.update(upd)
            cleaned = _normalize_row(row)
            if cleaned:
                # Preserve UI meta fields
                cleaned["idx"] = row.get("idx", idx)
                cleaned["source_excerpt"] = row.get("source_excerpt", "")
                rows[idx] = cleaned

    await db.ai_client_import_drafts.update_one(
        {"id": draft_id}, {"$set": {"rows": rows}},
    )
    return {"ok": True, "row": rows[idx]}


# ============================================================
# 4) Commit → insert clients
# ============================================================

class CommitRequest(BaseModel):
    min_confidence: int = Field(default=50, ge=0, le=100,
                                description="Skip rows below this confidence.")
    default_gdpr_consent: bool = Field(default=False)


@router.post("/draft/{draft_id}/commit")
async def commit_draft(
    draft_id: str,
    payload: CommitRequest,
    user: dict = Depends(require_roles("agency_admin", "agent", "super_admin")),
):
    db = Database.get()
    doc = await db.ai_client_import_drafts.find_one({"id": draft_id})
    if not doc:
        raise HTTPException(status_code=404, detail="draft_not_found")
    agency_id = await _agency_id(user)
    if doc.get("agency_id") != agency_id:
        raise HTTPException(status_code=403, detail="not_your_draft")
    if doc.get("status") == "committed":
        raise HTTPException(status_code=409, detail="draft_already_committed")

    rows = doc.get("rows") or []
    now = datetime.now(timezone.utc).isoformat()
    inserted: List[str] = []
    skipped: List[Dict[str, Any]] = []

    for r in rows:
        if r.get("_drop"):
            skipped.append({"idx": r.get("idx"), "reason": "dropped_by_user"})
            continue
        if int(r.get("confidence", 0)) < payload.min_confidence:
            skipped.append({"idx": r.get("idx"), "reason": "below_min_confidence"})
            continue
        client_doc = {
            "id": str(uuid4()),
            "agency_id": agency_id,
            "name": r["name"],
            "surname": r.get("surname"),
            "email": r.get("email"),
            "phone": r.get("phone"),
            "whatsapp": r.get("whatsapp"),
            "client_type": r.get("client_type") or "buyer",
            "status": r.get("status") or "new",
            "source": "ai_import",
            "preferences": r.get("preferences") or {},
            "notes": r.get("notes"),
            "gdpr_consent": payload.default_gdpr_consent,
            "created_at": now,
            "updated_at": now,
            "created_by": user.get("id") or "ai_import",
        }
        try:
            await db.clients.insert_one(client_doc)
            inserted.append(client_doc["id"])
        except Exception as e:
            skipped.append({"idx": r.get("idx"), "reason": f"db_error:{type(e).__name__}"})

    await db.ai_client_import_drafts.update_one(
        {"id": draft_id},
        {"$set": {"status": "committed", "committed_at": now, "committed_count": len(inserted)}},
    )
    return {
        "ok": True,
        "imported": len(inserted),
        "skipped": len(skipped),
        "skipped_details": skipped[:50],
        "total_rows": len(rows),
        "min_confidence_used": payload.min_confidence,
    }
